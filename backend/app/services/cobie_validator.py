"""
cobie_validator.py — Validare COBie XLSX și generare template.

Funcții:
- parse_cobie_xlsx: parsare XLSX → dict[sheet_name, rows]
- validate_structure: verificare sheet-uri + coloane obligatorii
- validate_against_project: verificare vs ProjectContext
- compute_cobie_score: calcul scor ponderat
- validate_cobie: orchestrator complet (cu DB)
- generate_cobie_template: generare XLSX template pre-populat
- get_cobie_validation_history / get_latest_cobie_validation: citire DB
"""

from __future__ import annotations

import datetime
import io
import json
import logging
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.sql_models import CobieValidationModel, KpiMeasurementModel
from app.repositories.projects_repository import (
    get_latest_project_context,
    get_project,
)
from app.schemas.cobie import (
    COBIE_REQUIRED_COLUMNS,
    COBIE_REQUIRED_SHEETS,
    CobieSheetCheck,
    CobieValidationCheck,
    CobieValidationResult,
)
from app.services.audit import log_action

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# Funcții pure (fără DB)
# ══════════════════════════════════════════════════════════════════════════════


class ParsedSheet:
    """Sheet parsat: headerele + rândurile de date."""
    __slots__ = ("headers", "rows")

    def __init__(self, headers: list[str], rows: list[dict]):
        self.headers = headers
        self.rows = rows


def parse_cobie_xlsx(file_path: str) -> dict[str, ParsedSheet]:
    """
    Parsează un fișier COBie XLSX.

    Returns:
        Dict {sheet_name: ParsedSheet(headers, rows)}
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    result: dict[str, ParsedSheet] = {}

    for sheet_name in wb.sheetnames:
        name = sheet_name.strip()
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[name] = ParsedSheet([], [])
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val
            if any(v is not None and str(v).strip() != "" for v in row_dict.values()):
                data_rows.append(row_dict)
        result[name] = ParsedSheet(headers, data_rows)

    wb.close()
    return result


def validate_structure(parsed: dict[str, ParsedSheet]) -> list[CobieSheetCheck]:
    """
    Validează structura COBie: sheet-uri prezente, coloane obligatorii, celule goale.
    """
    checks: list[CobieSheetCheck] = []
    parsed_lower = {k.lower(): (k, v) for k, v in parsed.items()}

    for sheet_name in COBIE_REQUIRED_SHEETS:
        expected_cols = COBIE_REQUIRED_COLUMNS.get(sheet_name, [])

        match = parsed_lower.get(sheet_name.lower())
        if not match:
            checks.append(CobieSheetCheck(
                sheet_name=sheet_name,
                status="missing",
                expected_columns=expected_cols,
                details=f"Sheet-ul '{sheet_name}' lipsește din fișier.",
            ))
            continue

        actual_name, sheet = match
        rows = sheet.rows
        # Use headers from ParsedSheet (works even for empty sheets)
        present_cols = sheet.headers if sheet.headers else (
            list(rows[0].keys()) if rows else []
        )
        present_lower = {c.lower() for c in present_cols}
        missing_cols = [c for c in expected_cols if c.lower() not in present_lower]

        # Count empty required cells
        check_cols = [c for c in expected_cols if c.lower() in present_lower]
        total_required = len(check_cols) * len(rows)
        empty_count = 0
        for row in rows:
            for col in check_cols:
                actual_col = next(
                    (k for k in row if k.lower() == col.lower()), None
                )
                if actual_col:
                    val = row.get(actual_col)
                    if val is None or str(val).strip() == "":
                        empty_count += 1

        # Determine status
        if missing_cols:
            status = "fail"
            details = f"Coloane lipsă: {', '.join(missing_cols)}"
        elif not rows:
            status = "warning"
            details = "Sheet gol (headere prezente, fără date)"
        elif total_required > 0:
            empty_pct = (empty_count / total_required) * 100 if total_required else 0
            if empty_pct > 30:
                status = "fail"
                details = f"{empty_pct:.0f}% celule obligatorii goale (>{30}% prag)"
            elif empty_pct > 10:
                status = "warning"
                details = f"{empty_pct:.0f}% celule obligatorii goale"
            else:
                status = "pass"
                details = f"{len(rows)} rânduri, structură completă"
        else:
            status = "pass"
            details = f"{len(rows)} rânduri"

        checks.append(CobieSheetCheck(
            sheet_name=sheet_name,
            status=status,
            row_count=len(rows),
            expected_columns=expected_cols,
            present_columns=present_cols,
            missing_columns=missing_cols,
            empty_required_cells=empty_count,
            total_required_cells=total_required,
            details=details,
        ))

    return checks


def validate_against_project(
    parsed: dict[str, ParsedSheet],
    context_json: dict | None,
) -> list[CobieValidationCheck]:
    """
    Validare COBie vs ProjectContext: project name, floors, disciplines, references.
    """
    checks: list[CobieValidationCheck] = []
    parsed_lower = {k.lower(): v.rows for k, v in parsed.items()}

    # 1. Facility.ProjectName matches project_name
    facility_rows = parsed_lower.get("facility", [])
    if facility_rows and context_json:
        project_name = context_json.get("project_name", "")
        found_name = None
        for row in facility_rows:
            for k, v in row.items():
                if k.lower() == "projectname" and v:
                    found_name = str(v).strip()
                    break
        if found_name and project_name:
            if found_name.lower() == project_name.lower():
                checks.append(CobieValidationCheck(
                    id="facility_project_name",
                    label="Facility.ProjectName = project_name",
                    status="pass",
                    details=f"'{found_name}' corespunde cu '{project_name}'",
                ))
            else:
                checks.append(CobieValidationCheck(
                    id="facility_project_name",
                    label="Facility.ProjectName = project_name",
                    status="warning",
                    details=f"'{found_name}' ≠ '{project_name}'",
                ))
        elif not found_name:
            checks.append(CobieValidationCheck(
                id="facility_project_name",
                label="Facility.ProjectName populat",
                status="fail",
                details="ProjectName lipsește din Facility",
            ))
    elif not facility_rows:
        checks.append(CobieValidationCheck(
            id="facility_project_name",
            label="Facility sheet prezent",
            status="fail",
            details="Sheet-ul Facility lipsește",
        ))

    # 2. Floor count > 0
    floor_rows = parsed_lower.get("floor", [])
    if floor_rows:
        checks.append(CobieValidationCheck(
            id="floor_count",
            label="Etaje definite",
            status="pass",
            details=f"{len(floor_rows)} etaj(e) definite",
        ))
    else:
        checks.append(CobieValidationCheck(
            id="floor_count",
            label="Etaje definite",
            status="fail",
            details="Niciun etaj definit în Floor",
        ))

    # 3. Discipline coverage (MEP → Systems with HVAC/Plumbing/Electrical)
    if context_json:
        disciplines = context_json.get("disciplines", [])
        system_rows = parsed_lower.get("system", [])
        mep_disciplines = {"mep", "hvac", "plumbing", "electrical"}
        has_mep = any(d.lower() in mep_disciplines for d in disciplines)

        if has_mep and system_rows:
            system_names = []
            for row in system_rows:
                for k, v in row.items():
                    if k.lower() == "name" and v:
                        system_names.append(str(v).lower())
            mep_keywords = {"hvac", "plumbing", "electrical", "mechanical", "ventilation"}
            found_mep = any(
                kw in name for name in system_names for kw in mep_keywords
            )
            if found_mep:
                checks.append(CobieValidationCheck(
                    id="mep_systems",
                    label="Sisteme MEP în COBie",
                    status="pass",
                    details=f"{len(system_rows)} sisteme definite, MEP acoperit",
                ))
            else:
                checks.append(CobieValidationCheck(
                    id="mep_systems",
                    label="Sisteme MEP în COBie",
                    status="warning",
                    details="Disciplina MEP definită în proiect dar sisteme specifice negăsite",
                ))
        elif has_mep:
            checks.append(CobieValidationCheck(
                id="mep_systems",
                label="Sisteme MEP în COBie",
                status="fail",
                details="Disciplina MEP definită în proiect dar sheet System gol",
            ))

    # 4. Component.Space references valid Space.Name
    component_rows = parsed_lower.get("component", [])
    space_rows = parsed_lower.get("space", [])
    if component_rows and space_rows:
        space_names = set()
        for row in space_rows:
            for k, v in row.items():
                if k.lower() == "name" and v:
                    space_names.add(str(v).strip().lower())

        invalid_refs = 0
        total_refs = 0
        for row in component_rows:
            for k, v in row.items():
                if k.lower() == "space" and v:
                    total_refs += 1
                    if str(v).strip().lower() not in space_names:
                        invalid_refs += 1

        if total_refs > 0:
            if invalid_refs == 0:
                checks.append(CobieValidationCheck(
                    id="component_space_ref",
                    label="Component.Space referințe valide",
                    status="pass",
                    details=f"Toate {total_refs} referințele Space sunt valide",
                ))
            else:
                checks.append(CobieValidationCheck(
                    id="component_space_ref",
                    label="Component.Space referințe valide",
                    status="warning",
                    details=f"{invalid_refs}/{total_refs} referințe Space invalide",
                ))

    # 5. Type with Category and Manufacturer populated
    type_rows = parsed_lower.get("type", [])
    if type_rows:
        missing_cat = 0
        missing_mfg = 0
        for row in type_rows:
            cat_val = next(
                (v for k, v in row.items() if k.lower() == "category"), None
            )
            mfg_val = next(
                (v for k, v in row.items() if k.lower() == "manufacturer"), None
            )
            if not cat_val or str(cat_val).strip() == "":
                missing_cat += 1
            if not mfg_val or str(mfg_val).strip() == "":
                missing_mfg += 1

        if missing_cat == 0 and missing_mfg == 0:
            checks.append(CobieValidationCheck(
                id="type_completeness",
                label="Type Category+Manufacturer complete",
                status="pass",
                details=f"{len(type_rows)} tipuri cu Category și Manufacturer",
            ))
        else:
            issues = []
            if missing_cat > 0:
                issues.append(f"{missing_cat} fără Category")
            if missing_mfg > 0:
                issues.append(f"{missing_mfg} fără Manufacturer")
            checks.append(CobieValidationCheck(
                id="type_completeness",
                label="Type Category+Manufacturer complete",
                status="warning",
                details=f"Din {len(type_rows)} tipuri: {', '.join(issues)}",
            ))

    return checks


def compute_cobie_score(
    sheet_checks: list[CobieSheetCheck],
    project_checks: list[CobieValidationCheck],
) -> tuple[float, str]:
    """
    Calculează scor COBie ponderat: structură 60%, project 40%.

    Returns:
        (score 0-100, overall_status)
    """
    # Structure score
    if sheet_checks:
        struct_points = 0
        for c in sheet_checks:
            if c.status == "pass":
                struct_points += 100
            elif c.status == "warning":
                struct_points += 50
        struct_score = struct_points / len(sheet_checks)
    else:
        struct_score = 0.0

    # Project checks score
    if project_checks:
        proj_points = 0
        for c in project_checks:
            if c.status == "pass":
                proj_points += 100
            elif c.status == "warning":
                proj_points += 50
        proj_score = proj_points / len(project_checks)
    else:
        proj_score = 0.0

    score = round(struct_score * 0.6 + proj_score * 0.4, 1)

    if score >= 80:
        overall_status = "pass"
    elif score >= 50:
        overall_status = "warning"
    else:
        overall_status = "fail"

    return score, overall_status


# ══════════════════════════════════════════════════════════════════════════════
# Funcții cu DB
# ══════════════════════════════════════════════════════════════════════════════


def validate_cobie(
    db: Session,
    project_id: int,
    file_path: str,
    filename: str,
    file_size_bytes: int | None = None,
    validation_type: str = "full",
) -> CobieValidationResult:
    """
    Orchestrator: parse → validate → score → save DB.
    """
    # 1. Parse
    parsed = parse_cobie_xlsx(file_path)

    # 2. Validate structure
    sheet_checks = validate_structure(parsed)

    # 3. Validate against project context
    project_checks: list[CobieValidationCheck] = []
    ctx_entry = get_latest_project_context(db, project_id)
    context_json = ctx_entry.context_json if ctx_entry else None

    if validation_type == "full" and context_json:
        project_checks = validate_against_project(parsed, context_json)

    # 4. Compute score
    score, overall_status = compute_cobie_score(sheet_checks, project_checks)

    # 5. Count statuses
    all_statuses = (
        [c.status for c in sheet_checks]
        + [c.status for c in project_checks]
    )
    pass_count = sum(1 for s in all_statuses if s == "pass")
    warning_count = sum(1 for s in all_statuses if s == "warning")
    fail_count = sum(1 for s in all_statuses if s in ("fail", "missing"))

    # 6. Recommendations
    recommendations: list[str] = []
    missing_sheets = [c.sheet_name for c in sheet_checks if c.status == "missing"]
    if missing_sheets:
        recommendations.append(
            f"Adaugă sheet-urile lipsă: {', '.join(missing_sheets)}"
        )
    failed_sheets = [c.sheet_name for c in sheet_checks if c.status == "fail"]
    if failed_sheets:
        recommendations.append(
            f"Corectează sheet-urile cu erori: {', '.join(failed_sheets)}"
        )
    for c in project_checks:
        if c.status == "fail":
            recommendations.append(f"{c.label}: {c.details}")
    if not context_json:
        recommendations.append(
            "Completează fișa proiectului (ProjectContext) pentru validare completă."
        )

    result = CobieValidationResult(
        score=score,
        overall_status=overall_status,
        total_checks=len(all_statuses),
        pass_count=pass_count,
        warning_count=warning_count,
        fail_count=fail_count,
        sheet_checks=sheet_checks,
        project_checks=project_checks,
        recommendations=recommendations,
    )

    # 7. Build sheet stats
    sheet_stats = {}
    for c in sheet_checks:
        sheet_stats[c.sheet_name] = {
            "status": c.status,
            "row_count": c.row_count,
            "missing_columns": c.missing_columns,
        }

    # 8. Save to DB
    validation = CobieValidationModel(
        project_id=project_id,
        filename=filename,
        file_path=file_path,
        file_size_bytes=file_size_bytes,
        validation_type=validation_type,
        overall_status=overall_status,
        score=score,
        total_checks=len(all_statuses),
        pass_count=pass_count,
        warning_count=warning_count,
        fail_count=fail_count,
        results_json=result.model_dump(),
        sheet_stats_json=sheet_stats,
    )
    db.add(validation)
    db.flush()

    # 9. Register KPI
    kpi = KpiMeasurementModel(
        project_id=project_id,
        kpi_name="cobie_completeness",
        category="data_quality",
        value=score,
        target_value=80.0,
        measurement_date=datetime.date.today(),
    )
    db.add(kpi)

    # 10. Audit log
    log_action(db, project_id, "validate_cobie", {
        "filename": filename,
        "score": score,
        "overall_status": overall_status,
        "total_checks": len(all_statuses),
    })

    db.commit()
    return result


def generate_cobie_template(
    db: Session,
    project_id: int,
    include_ai_suggestions: bool = False,
    target_sheets: list[str] | None = None,
) -> io.BytesIO:
    """
    Generează template COBie XLSX cu headers stilizate.

    Pre-populează Facility cu date din ProjectContext dacă disponibil.
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheets_to_create = target_sheets or COBIE_REQUIRED_SHEETS

    # Get project context for pre-population
    ctx_entry = get_latest_project_context(db, project_id)
    context_json = ctx_entry.context_json if ctx_entry else None
    project = get_project(db, project_id)

    for sheet_name in sheets_to_create:
        if sheet_name not in COBIE_REQUIRED_COLUMNS:
            continue

        ws = wb.create_sheet(title=sheet_name)
        columns = COBIE_REQUIRED_COLUMNS[sheet_name]

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Set column widths
        for col_idx in range(1, len(columns) + 1):
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = 18

        # Freeze panes (header row)
        ws.freeze_panes = "A2"

        # Pre-populate Facility from ProjectContext
        if sheet_name == "Facility" and context_json:
            row_data = {
                "Name": context_json.get("project_name", ""),
                "CreatedBy": "Agent BIM",
                "CreatedOn": datetime.date.today().isoformat(),
                "Category": context_json.get("project_type", ""),
                "ProjectName": context_json.get("project_name", ""),
                "SiteName": context_json.get("site_name", ""),
                "LinearUnits": "meters",
                "AreaUnits": "square meters",
                "VolumeUnits": "cubic meters",
                "CurrencyUnit": "RON",
                "AreaMeasurement": "Gross Area",
            }
            for col_idx, col_name in enumerate(columns, 1):
                val = row_data.get(col_name, "")
                if val:
                    ws.cell(row=2, column=col_idx, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    if project:
        log_action(db, project_id, "generate_cobie_template", {
            "sheets": sheets_to_create,
            "include_ai_suggestions": include_ai_suggestions,
        })

    return output


def get_cobie_validation_history(
    db: Session, project_id: int
) -> list[CobieValidationModel]:
    """Returnează istoricul validărilor COBie (cele mai recente primele)."""
    return (
        db.query(CobieValidationModel)
        .filter(CobieValidationModel.project_id == project_id)
        .order_by(CobieValidationModel.created_at.desc())
        .all()
    )


def get_latest_cobie_validation(
    db: Session, project_id: int
) -> CobieValidationModel | None:
    """Returnează ultima validare COBie."""
    return (
        db.query(CobieValidationModel)
        .filter(CobieValidationModel.project_id == project_id)
        .order_by(CobieValidationModel.created_at.desc())
        .first()
    )
